import pandas as pd
import numpy as np
from datetime import date, datetime, timedelta
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, text, or_, and_, select
from models.restock import InventorySnapshot, InboundShipmentDetail, ReplenishmentDecision
from database.database import engine
from utils.excel_reader import safe_read_excel

# Excel中文名 -> 数据库字段名 映射
# 注意：采购配置字段（purchase_plan_days, purchase_lead_time, qc_days 等）不导入数据库
FIELD_MAPPING = {
    "欧洲/北美汇总行": "summary_flag",
    "ASIN": "asin",
    "父ASIN": "parent_asin",
    "MSKU": "msku",
    "FNSKU": "fnsku",
    "SKU": "sku",
    "品名": "product_name",
    "标题": "title",
    "店铺": "account",
    "国家（地区）": "country",
    "分类": "category",
    "品牌": "brand",
    "补货状态": "replenishment_status",
    # 采购配置字段不导入（采购计划天数、采购交期、质检天数、海外仓至FBA天数、安全天数、采购频率、发货频率、备货时长）
    "3天销量": "sales_3d",
    "7天销量": "sales_7d",
    "14天销量": "sales_14d",
    "30天销量": "sales_30d",
    "60天销量": "sales_60d",
    "90天销量": "sales_90d",
    # 日均销量字段不导入，由公式计算得出
    "3天日均": "daily_avg_3d",
    "7天日均": "daily_avg_7d",
    "14天日均": "daily_avg_14d",
    "30天日均": "daily_avg_30d",
    "60天日均": "daily_avg_60d",
    "90天日均": "daily_avg_90d",
    # daily_sales 不导入，由公式计算得出（不要加在FIELD_MAPPING中）
    "可售天数(总)": "days_supply_total",
    "可售天数(FBA)": "days_supply_fba",
    "可售天数(FBA + 在途)": "days_supply_fba_inbound",
    "断货时间": "stockout_date",
    # "日均销量": "daily_sales",  # 由公式计算，不直接导入
    "销量预测": "sales_forecast",
    "FBA库存": "fba_stock",
    "FBA在途": "fba_inbound",
    "FBA在途详情": "fba_inbound_detail",
    "可售": "fba_available",
    "待调仓": "fba_pending_transfer",
    "调仓中": "fba_in_transfer",
    "入库中": "fba_inbound_processing",
    "本地可用": "local_available",
    "总库存": "total_stock",
    "3个月库龄": "age_0_3",
    "3-6个月库龄": "age_3_6",
    "6-9个月库龄": "age_6_9",
    "9-12个月库龄": "age_9_12",
    "12个月以上库龄": "age_12_plus",
    "毛利率": "gross_margin",
}

NUMERIC_FIELDS = [
    # 销量数据
    "sales_3d", "sales_7d", "sales_14d", "sales_30d", "sales_60d", "sales_90d",
    "daily_avg_3d", "daily_avg_7d", "daily_avg_14d", "daily_avg_30d",
    "daily_avg_60d", "daily_avg_90d", "daily_sales", "sales_forecast",
    # 可售天数
    "days_supply_total", "days_supply_fba", "days_supply_fba_inbound",
    # 库存数据
    "fba_stock", "fba_inbound", "fba_available", "fba_pending_transfer",
    "fba_in_transfer", "fba_inbound_processing", "local_available", "total_stock",
    # 库龄分布
    "age_0_3", "age_3_6", "age_6_9", "age_9_12", "age_12_plus",
]

DATE_FIELDS = ["stockout_date"]
LEAD_TIME = 100

# 默认日均销量公式配置
DEFAULT_DAILY_SALES_WEIGHTS = {
    "daily_avg_3d": 0.0,
    "daily_avg_7d": 0.2,
    "daily_avg_14d": 0.2,
    "daily_avg_30d": 0.2,
    "daily_avg_60d": 0.2,
    "daily_avg_90d": 0.2,
}


def _build_account_match_conditions(account_value: str):
    """构建 account 字段匹配条件

    精确匹配普通行 + 边界匹配汇总行（汇总行用 '、' 分隔多个店铺）。
    避免使用 like('%val%') 导致 'G-UK' 误匹配 'Roaring-UK-UK'
    （MySQL LIKE 默认不区分大小写，'%G-UK%' 会匹配 'RoarinG-UK-UK'）。
    """
    return [
        InventorySnapshot.account == account_value,
        InventorySnapshot.account.like(f"{account_value}、%"),
        InventorySnapshot.account.like(f"%、{account_value}、%"),
        InventorySnapshot.account.like(f"%、{account_value}"),
    ]


def _build_account_filter(account_values):
    """为多个 account 值构建 or_ 过滤条件（支持 list 和 str）"""
    conditions = []
    if isinstance(account_values, list):
        for a in account_values:
            conditions.extend(_build_account_match_conditions(a))
    elif isinstance(account_values, str):
        conditions.extend(_build_account_match_conditions(account_values))
    return conditions


def _get_daily_sales_weights(db) -> dict:
    """从数据库获取日均销量公式配置"""
    try:
        from models.business_settings import BusinessSettings
        import json

        setting = db.query(BusinessSettings).filter(
            BusinessSettings.setting_type == "daily_sales"
        ).first()

        if setting and setting.formula_config:
            config = json.loads(setting.formula_config)
            weights = {}
            for w in config.get("weights", []):
                period = w.get("period", "")
                weight = w.get("weight", 0)
                # 将 period 转换为字段名
                field_map = {
                    "3d": "daily_avg_3d",
                    "7d": "daily_avg_7d",
                    "14d": "daily_avg_14d",
                    "30d": "daily_avg_30d",
                    "60d": "daily_avg_60d",
                    "90d": "daily_avg_90d",
                }
                if period in field_map:
                    weights[field_map[period]] = weight
            return weights
    except Exception as e:
        print(f"获取日均销量公式配置失败: {e}")

    return DEFAULT_DAILY_SALES_WEIGHTS


def _calculate_daily_sales(df: pd.DataFrame, db=None) -> np.ndarray:
    """
    根据公式计算日均销量
    公式：daily_sales = daily_avg_3d * w1 + daily_avg_7d * w2 + ...
    """
    weights = DEFAULT_DAILY_SALES_WEIGHTS
    if db is not None:
        try:
            weights = _get_daily_sales_weights(db)
        except Exception as e:
            print(f"[WARN] 获取公式配置失败，使用默认: {e}")

    print(f"[DEBUG] 计算日均销量，权重: {weights}")

    n = len(df)
    result = np.zeros(n)

    for field, weight in weights.items():
        if field in df.columns:
            values = pd.to_numeric(df[field], errors="coerce").fillna(0).values
            result += values * weight

    print(f"[DEBUG] 计算完成: {n} 条记录, 平均日均销量={np.mean(result):.2f}")
    return result


def _parse_inbound_details_fast(raw_text: str) -> list:
    """快速解析在途详情"""
    if not raw_text or pd.isna(raw_text) or not str(raw_text).strip():
        return []
    
    results = []
    header_keywords = {"货件单号", "shipment id", "shipmentid", "shipment_id", "shipment", "单号", "id"}
    
    for line in str(raw_text).strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        parts = line.split("|")
        if not parts:
            continue
        
        shipment_id = parts[0].strip()
        if not shipment_id:
            continue
        if any(kw in shipment_id.lower() for kw in header_keywords):
            continue
        if not any(c.isdigit() for c in shipment_id):
            continue
        
        quantity = None
        estimated_date = None
        logistics_method = None
        transport_method = None
        ship_date = None

        try:
            if len(parts) >= 3 and parts[2].strip():
                quantity = int(float(parts[2].strip()))
        except:
            pass
        if len(parts) >= 4 and parts[3].strip():
            logistics_method = parts[3].strip()
        if len(parts) >= 5 and parts[4].strip():
            transport_method = parts[4].strip()
        try:
            if len(parts) >= 6 and parts[5].strip():
                ship_date = datetime.strptime(parts[5].strip(), "%Y-%m-%d").date()
        except:
            pass
        try:
            if len(parts) >= 7 and parts[6].strip():
                estimated_date = datetime.strptime(parts[6].strip(), "%Y-%m-%d").date()
            elif parts[-1].strip() and len(parts[-1]) == 10:
                estimated_date = datetime.strptime(parts[-1].strip(), "%Y-%m-%d").date()
        except:
            pass

        results.append({
            "shipment_id": shipment_id,
            "quantity": quantity,
            "logistics_method": logistics_method,
            "transport_method": transport_method,
            "ship_date": ship_date,
            "estimated_available_date": estimated_date,
            "raw_text": line,
        })
    return results


def import_inventory_data(db: Session, file_path: str = None, file_content: bytes = None, filename: str = None) -> dict:
    """导入库存Excel数据 - 极速版"""
    import io
    import time as _time
    
    _t_start = _time.time()
    
    def _log(msg):
        elapsed = _time.time() - _t_start
        print(f"[导入 {elapsed:>6.1f}s] {msg}")
    
    # ========== 步骤1: 读取Excel ==========
    _log(f"开始读取Excel: {filename or file_path or '上传文件'}")
    if file_content:
        df = safe_read_excel(file_content)
    elif file_path:
        df = pd.read_excel(file_path)
    else:
        raise ValueError("请提供 file_path 或 file_content")

    today = date.today()
    total_rows = len(df)
    _log(f"Excel读取完成: {total_rows} 行, {len(df.columns)} 列")
    _log(f"列名: {list(df.columns[:10])}{'...' if len(df.columns) > 10 else ''}")

    # 重命名列
    df = df.rename(columns={k: v for k, v in FIELD_MAPPING.items() if k in df.columns})

    # 数据清洗
    _log("数据清洗中...")
    if "summary_flag" in df.columns:
        df["summary_flag"] = df["summary_flag"].fillna("0").apply(
            lambda x: "0" if str(x).strip() in ("", "0") else str(x).strip()
        )
    
    for col in NUMERIC_FIELDS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    str_cols = ["asin", "parent_asin", "msku", "fnsku", "sku", "product_name", "title", 
                "account", "country", "category", "brand", "fba_inbound_detail"]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)

    df["snapshot_date"] = today
    df["tenant_id"] = 1
    _log("数据清洗完成")

    # ========== 步骤2: 建立旧数据索引（按产品取最新的一条） ==========
    _log("加载现有数据索引...")
    try:
        existing_rows = db.execute(text("""
            SELECT id, asin, account, country, summary_flag, snapshot_date
            FROM inventory_snapshots
            ORDER BY snapshot_date DESC
        """)).fetchall()
    except Exception:
        existing_rows = []

    idx_normal = {}   # 普通行: (asin, account, country) → (id, snapshot_date)
    idx_summary = {}  # 汇总行: (summary_flag, asin, country) → (id, snapshot_date)

    for row in existing_rows:
        sf = row.summary_flag or ""
        if sf == "是":
            # 汇总行只进 idx_summary，避免孤儿清理时用 (asin, account, country) 误判
            key_s = (sf, row.asin or "", row.country or "")
            if key_s not in idx_summary:
                idx_summary[key_s] = (row.id, row.snapshot_date)
        else:
            key_n = (row.asin or "", row.account or "", row.country or "")
            if key_n not in idx_normal:
                idx_normal[key_n] = (row.id, row.snapshot_date)

    _log(f"现有数据索引加载完成: {len(existing_rows)} 条原始, {len(idx_normal)} 个独立产品")

    # ========== 步骤3: UPSERT 快照 ==========
    _log(f"开始 UPSERT inventory_snapshots ({total_rows} 条)...")
    available_columns = [col for col in FIELD_MAPPING.values() if col in df.columns] + ["snapshot_date", "tenant_id"]
    df_to_insert = df[available_columns].copy()
    df_to_insert["stock_up_duration"] = LEAD_TIME

    _log("计算日均销量...")
    daily_sales_values = _calculate_daily_sales(df, db)
    df_to_insert["daily_sales"] = daily_sales_values

    columns = list(df_to_insert.columns)
    batch_size = 1000
    snapshot_ids = []
    inserted_count = 0
    updated_count = 0

    for i in range(0, len(df_to_insert), batch_size):
        batch = df_to_insert.iloc[i:i+batch_size]
        update_batch = []
        insert_batch = []
        batch_ids = []
        
        for _, row in batch.iterrows():
            row_summary_flag = str(row.get("summary_flag", "0"))
            row_asin = str(row.get("asin", ""))
            row_account = str(row.get("account", ""))
            row_country = str(row.get("country", ""))
            
            if row_summary_flag == "是":
                match_key = (row_summary_flag, row_asin, row_country)
                existing = idx_summary.get(match_key)
            else:
                match_key = (row_asin, row_account, row_country)
                existing = idx_normal.get(match_key)
            
            if existing is not None:
                existing_id = existing[0]  # 从 (id, snapshot_date) 元组中取 id
                set_parts = []
                for col in columns:
                    val = row[col]
                    if pd.isna(val):
                        continue  # 跳过空值，保留数据库原值
                    elif isinstance(val, str):
                        safe = str(val).replace("'", "''")
                        set_parts.append(f"{col}='{safe}'")
                    elif isinstance(val, (int, float)):
                        set_parts.append(f"{col}={val}")
                    elif isinstance(val, date):
                        set_parts.append(f"{col}='{val}'")
                    else:
                        safe = str(val).replace("'", "''")
                        set_parts.append(f"{col}='{safe}'")
                update_batch.append(f"UPDATE inventory_snapshots SET deleted_at=NULL,updated_at=NOW(),{','.join(set_parts)} WHERE id={existing_id}")
                batch_ids.append(existing_id)
                updated_count += 1
            else:
                row_values = []
                for col in columns:
                    val = row[col]
                    if pd.isna(val):
                        row_values.append("NULL")
                    elif isinstance(val, str):
                        row_values.append(f"'{str(val).replace(chr(39), chr(39)+chr(39))}'")
                    elif isinstance(val, (int, float)):
                        row_values.append(str(val))
                    elif isinstance(val, date):
                        row_values.append(f"'{val}'")
                    else:
                        row_values.append(f"'{str(val)}'")
                insert_batch.append(f"({','.join(row_values)})")
                batch_ids.append(None)
                inserted_count += 1
        
        for sql in update_batch:
            db.execute(text(sql))
        db.commit()
        
        if insert_batch:
            sql = f"INSERT INTO inventory_snapshots ({','.join(columns)}) VALUES {','.join(insert_batch)}"
            db.execute(text(sql))
            db.commit()
            result = db.execute(text("SELECT LAST_INSERT_ID()")).scalar()
            if result:
                last_id = result
                for j in range(len(batch_ids) - 1, -1, -1):
                    if batch_ids[j] is None:
                        batch_ids[j] = last_id
                        last_id -= 1
        
        snapshot_ids.extend(batch_ids)
        
        done = min(i + batch_size, len(df_to_insert))
        if done % 5000 == 0 or done >= len(df_to_insert):
            _log(f"  UPSERT进度: {done}/{len(df_to_insert)}, 更新{updated_count} 新增{inserted_count}")

    _log(f"inventory_snapshots UPSERT完成: {total_rows} 条 (更新{updated_count} 新增{inserted_count})")

    # ========== 步骤4: 确认快照ID ==========
    _log(f"确认快照ID: {len(snapshot_ids)} 个")
    assert len(snapshot_ids) == total_rows, f"snapshot_ids 数量 ({len(snapshot_ids)}) 与 Excel 行数 ({total_rows}) 不匹配"

    # ========== 步骤5: 解析在途详情 ==========
    _log("开始解析在途详情...")
    inbound_records = []
    df_reset = df.reset_index(drop=True)
    
    # 先统计有多少条有在途详情
    detail_count = 0
    for idx in range(len(df_reset)):
        raw_detail = df_reset.iloc[idx].get("fba_inbound_detail", "")
        if raw_detail and str(raw_detail).strip() and str(raw_detail).strip() != "nan":
            detail_count += 1
    _log(f"有在途详情的记录: {detail_count}/{total_rows} 条")
    
    # 解析
    for idx in range(len(df_reset)):
        row = df_reset.iloc[idx]
        raw_detail = row.get("fba_inbound_detail", "")
        if raw_detail and str(raw_detail).strip() and str(raw_detail).strip() != "nan":
            if idx < len(snapshot_ids):
                details = _parse_inbound_details_fast(raw_detail)
                for d in details:
                    inbound_records.append({
                        "tenant_id": 1,
                        "snapshot_id": snapshot_ids[idx],
                        "asin": row.get("asin", ""),
                        "account": row.get("account", ""),
                        "country": row.get("country", ""),
                        **d
                    })
        # 每10000行输出进度
        if (idx + 1) % 10000 == 0:
            _log(f"  解析进度: {idx+1}/{total_rows}, 在途详情: {len(inbound_records)} 条")

    _log(f"在途详情解析完成: {len(inbound_records)} 条")

    # ========== 步骤6: 导入在途详情 (UPSERT) ==========
    if inbound_records:
        _log(f"开始 UPSERT inbound_shipment_details ({len(inbound_records)} 条)...")
        
        # 加载当前批次涉及的在途详情现有记录（含软删），建立索引
        all_sids = set(r['snapshot_id'] for r in inbound_records)
        ids_str = ",".join(str(sid) for sid in all_sids)
        existing_inbound = db.execute(text(f"""
            SELECT id, snapshot_id, shipment_id, deleted_at
            FROM inbound_shipment_details
            WHERE snapshot_id IN ({ids_str})
        """)).fetchall()
        inbound_idx = {}
        inbound_deleted_idx = {}
        for row in existing_inbound:
            key = (row.snapshot_id, row.shipment_id)
            if row.deleted_at is None:
                inbound_idx[key] = row.id
            elif key not in inbound_idx:
                inbound_deleted_idx[key] = row.id
        
        matched_inbound_keys = set()
        batch_size = 1000
        updated_inbound = 0
        inserted_inbound = 0
        
        for i in range(0, len(inbound_records), batch_size):
            batch = inbound_records[i:i+batch_size]
            update_batch = []
            insert_values = []
            
            for r in batch:
                key = (r['snapshot_id'], r['shipment_id'])
                existing_id = inbound_idx.get(key)
                is_deleted_restore = False
                if existing_id is None:
                    existing_id = inbound_deleted_idx.get(key)
                    if existing_id is not None:
                        is_deleted_restore = True
                
                if existing_id is not None:
                    matched_inbound_keys.add(key)
                    # UPDATE — 跳过空值
                    set_parts = []
                    sid = r['snapshot_id']
                    asin = str(r.get('asin', '')).replace("'", "''")
                    account = str(r.get('account', '')).replace("'", "''")
                    country = str(r.get('country', '')).replace("'", "''")
                    shipment_id = str(r.get('shipment_id', '')).replace("'", "''")
                    if asin:
                        set_parts.append(f"asin='{asin}'")
                    if account:
                        set_parts.append(f"account='{account}'")
                    if country:
                        set_parts.append(f"country='{country}'")
                    if shipment_id:
                        set_parts.append(f"shipment_id='{shipment_id}'")
                    qty = r.get('quantity', 0) or 0
                    set_parts.append(f"quantity={qty}")
                    lm = str(r.get('logistics_method', '') or '')
                    if lm:
                        set_parts.append(f"logistics_method='{lm.replace(chr(39), chr(39)+chr(39))}'")
                    tm = str(r.get('transport_method', '') or '')
                    if tm:
                        set_parts.append(f"transport_method='{tm.replace(chr(39), chr(39)+chr(39))}'")
                    if r.get('ship_date'):
                        set_parts.append(f"ship_date='{r['ship_date']}'")
                    if r.get('estimated_available_date'):
                        set_parts.append(f"estimated_available_date='{r['estimated_available_date']}'")
                    raw = str(r.get('raw_text', '')).replace("'", "''")
                    if raw:
                        set_parts.append(f"raw_text='{raw}'")
                    
                    if set_parts:
                        if is_deleted_restore:
                            set_parts.insert(0, "deleted_at=NULL")
                        update_batch.append(f"UPDATE inbound_shipment_details SET {','.join(set_parts)} WHERE id={existing_id}")
                        updated_inbound += 1
                else:
                    # INSERT
                    sid = r['snapshot_id']
                    a = str(r.get('asin', '')).replace("'", "''")
                    acc = str(r.get('account', '')).replace("'", "''")
                    cntry = str(r.get('country', '')).replace("'", "''")
                    ship = str(r.get('shipment_id', '')).replace("'", "''")
                    qty = r.get('quantity', 0) or 0
                    lm = str(r.get('logistics_method', '') or '').replace("'", "''")
                    tm = str(r.get('transport_method', '') or '').replace("'", "''")
                    sd = f"'{r['ship_date']}'" if r.get('ship_date') else "NULL"
                    ed = f"'{r['estimated_available_date']}'" if r.get('estimated_available_date') else "NULL"
                    raw = str(r.get('raw_text', '')).replace("'", "''")
                    insert_values.append(f"({r['tenant_id']}, {sid}, '{a}', '{acc}', '{cntry}', '{ship}', {qty}, '{lm}', '{tm}', {sd}, {ed}, '{raw}')")
                    inserted_inbound += 1
            
            for sql in update_batch:
                db.execute(text(sql))
            db.commit()
            
            if insert_values:
                sql = f"""
                    INSERT INTO inbound_shipment_details
                    (tenant_id, snapshot_id, asin, account, country, shipment_id, quantity, logistics_method, transport_method, ship_date, estimated_available_date, raw_text)
                    VALUES {','.join(insert_values)}
                """
                db.execute(text(sql))
                db.commit()
            
            done = min(i + batch_size, len(inbound_records))
            if done % 5000 == 0 or done >= len(inbound_records):
                _log(f"  UPSERT进度: {done}/{len(inbound_records)}, 更新{updated_inbound} 新增{inserted_inbound}")
        
        # 孤儿清理：DB有但Excel没有 → 软删除
        orphan_inbound_ids = []
        for key, eid in inbound_idx.items():
            if key not in matched_inbound_keys:
                orphan_inbound_ids.append(eid)
        if orphan_inbound_ids:
            ids_str = ",".join(str(i) for i in orphan_inbound_ids)
            db.execute(text(f"UPDATE inbound_shipment_details SET deleted_at = NOW() WHERE id IN ({ids_str})"))
            db.commit()
            _log(f"  孤儿在途详情软删除: {len(orphan_inbound_ids)} 条")
        
        # fba_inbound 校验修正
        _log("fba_inbound 校验修正...")
        for sid in all_sids:
            sum_qty = db.execute(text(
                f"SELECT COALESCE(SUM(quantity), 0) FROM inbound_shipment_details WHERE snapshot_id = {sid} AND deleted_at IS NULL"
            )).scalar() or 0
            old_val = db.execute(text(
                f"SELECT fba_inbound FROM inventory_snapshots WHERE id = {sid} AND deleted_at IS NULL"
            )).scalar()
            if old_val is not None and old_val != sum_qty:
                db.execute(text(
                    f"UPDATE inventory_snapshots SET fba_inbound = {sum_qty} WHERE id = {sid} AND deleted_at IS NULL"
                ))
                db.commit()
                _log(f"  fba_inbound 校验修正: snapshot_id={sid}, {old_val} → {sum_qty}")
        
        _log(f"inbound_shipment_details UPSERT完成: 更新{updated_inbound} 新增{inserted_inbound}")
    else:
        _log("没有在途详情需要导入")

    # ========== 步骤7: 孤儿清理（软删除） ==========
    _log("孤儿清理...")
    orphan_ids = set()
    imported_normal_keys = {(str(r.get("asin","")), str(r.get("account","")), str(r.get("country","")))
                            for _, r in df.iterrows() if str(r.get("summary_flag","")) != "是"}
    imported_summary_keys = {("是", str(r.get("asin","")), str(r.get("country","")))
                             for _, r in df.iterrows() if str(r.get("summary_flag","")) == "是"}
    # 普通行：DB 有但 Excel 没有 → 孤儿
    for key, val in idx_normal.items():
        if key not in imported_normal_keys:
            orphan_ids.add(val[0])
    # 汇总行：DB 有但 Excel 没有 → 孤儿
    for key, val in idx_summary.items():
        if key not in imported_summary_keys:
            orphan_ids.add(val[0])

    if orphan_ids:
        ids_str = ",".join(str(i) for i in orphan_ids)
        db.execute(text(f"UPDATE replenishment_decisions SET deleted_at = NOW() WHERE snapshot_id IN ({ids_str})"))
        db.execute(text(f"UPDATE inbound_shipment_details SET deleted_at = NOW() WHERE snapshot_id IN ({ids_str})"))
        db.execute(text(f"UPDATE inventory_snapshots SET deleted_at = NOW() WHERE id IN ({ids_str})"))
        db.commit()
        _log(f"孤儿清理完成: 软删除 {len(orphan_ids)} 条")
    else:
        _log("无孤儿数据需要清理")

    # ========== 步骤8: 计算补货决策 ==========
    _log("开始计算补货决策 replenishment_decisions (UPSERT)...")
    calc_result = _calculate_replenishment_fast(db, df, snapshot_ids, today)
    _log(f"replenishment_decisions 计算完成: {calc_result}")

    # ========== 完成 ==========
    total_time = _time.time() - _t_start
    _log(f"全部完成! 总耗时 {total_time:.1f}s")
    print(f"\n{'='*60}")
    print(f"导入汇总:")
    print(f"  inventory_snapshots:      {total_rows} 条")
    print(f"  inbound_shipment_details: {len(inbound_records)} 条")
    print(f"  replenishment_decisions:  {calc_result.get('total', 'N/A')} 条")
    print(f"  总耗时: {total_time:.1f}s")
    print(f"{'='*60}")

    return {
        "total_rows": total_rows,
        "imported": total_rows,
        "inbound_details": len(inbound_records),
        "snapshot_date": today.isoformat(),
        "calculation": calc_result,
    }


def _calculate_replenishment_fast(db: Session, df: pd.DataFrame, snapshot_ids: list, target_date: date, progress_callback=None) -> dict:
    """批量补货计算 - DataFrame向量化（含本地仓库存）"""

    # 准备计算数据
    n = len(df)
    summary_flags = df["summary_flag"].fillna("0").values if "summary_flag" in df.columns else np.array(["0"] * n)
    daily_sales = df["daily_sales"].fillna(0).values if "daily_sales" in df.columns else np.zeros(n)
    fba_stock = df["fba_stock"].fillna(0).values if "fba_stock" in df.columns else np.zeros(n)
    fba_inbound = df["fba_inbound"].fillna(0).values if "fba_inbound" in df.columns else np.zeros(n)
    local_inventory = df["local_inventory"].fillna(0).values if "local_inventory" in df.columns else np.zeros(n)
    inspection_qty = df["inspection_quantity"].fillna(0).values if "inspection_quantity" in df.columns else np.zeros(n)
    # 动态计算总库存：fba_stock + fba_inbound + local_inventory - inspection_quantity
    total_stock = fba_stock + fba_inbound + local_inventory - inspection_qty
    asins = df["asin"].fillna("").values if "asin" in df.columns else np.array([""] * n)
    skus = df["sku"].fillna("").values if "sku" in df.columns else np.array([""] * n)
    accounts = df["account"].fillna("").values if "account" in df.columns else np.array([""] * n)
    countries = df["country"].fillna("").values if "country" in df.columns else np.array([""] * n)

    # 本地仓库存直接使用 inventory_snapshots.local_inventory 字段值
    effective_local = local_inventory

    if progress_callback:
        progress_callback("计算中", 0, 100)

    # 向量化计算：future_stock = FBA总库存 + FBA在途 + 本地仓库存 - 查验货件
    # 查验货件尚未入库 FBA，不应计入未来可用库存，否则会低估补货需求
    future_stock = fba_stock + fba_inbound + effective_local - inspection_qty
    days_of_supply = np.where(
        (total_stock > 0) & (daily_sales > 0),
        np.minimum(total_stock / daily_sales, 365),
        365
    )
    demand = daily_sales * LEAD_TIME
    suggest_qty = np.maximum(0, demand - future_stock)
    # 向上取整到50的倍数（0保持0）
    suggest_qty = np.where(suggest_qty > 0, np.ceil(suggest_qty / 50) * 50, 0)
    stockout_days = np.maximum(0, LEAD_TIME - days_of_supply).astype(int)

    if progress_callback:
        progress_callback("计算中", 30, 100)

    # 风险等级
    risk_levels = np.where(days_of_supply <= 30, "红", np.where(days_of_supply <= 60, "黄", "绿"))

    # 断货日期
    stockout_dates = []
    for d in days_of_supply:
        if d >= 365:
            stockout_dates.append("-")
        else:
            stockout_dates.append(target_date + timedelta(days=int(d)))

    # 原因（含本地仓库存信息）
    reasons = np.where(
        days_of_supply < LEAD_TIME,
        [f"可售{round(d,1)}天，低于{LEAD_TIME}天备货周期，建议补货{int(s)}件"
         + (f"（含本地仓{int(lq)}件）" if lq > 0 else "")
         for d, s, lq in zip(days_of_supply, suggest_qty, effective_local)],
        [f"可售{round(d,1)}天，超过{LEAD_TIME}天备货周期，库存充足"
         + (f"（含本地仓{int(lq)}件）" if lq > 0 else "")
         for d, lq in zip(days_of_supply, effective_local)]
    )

    # 共享库存处理
    shared_mask = summary_flags == "共享库存"
    demand[shared_mask] = 0
    suggest_qty[shared_mask] = 0
    risk_levels[shared_mask] = "绿"
    reasons[shared_mask] = "共享库存子行，库存由汇总行统一管理"

    # 极低销量处理
    low_sales_mask = daily_sales <= 0.1
    suggest_qty[low_sales_mask] = 0
    risk_levels[low_sales_mask] = "绿"
    reasons[low_sales_mask] = "日均销量极低（≤0.1），当前库存充足，无需补货"

    if progress_callback:
        progress_callback("写入数据库", 40, 100)

    # ========== Diff 校验 + 批量 UPSERT（INSERT ... ON DUPLICATE KEY UPDATE） ==========

    # 1. 读取现有决策的全部字段值（用于 diff 校验）
    existing_sids = set(sid for sid in snapshot_ids if sid is not None)
    existing_values = {}
    if existing_sids:
        try:
            existing_rows = db.execute(text(f"""
                SELECT snapshot_id, summary_flag, asin, sku, account, country, snapshot_date,
                       future_stock, demand, suggest_qty, days_of_supply,
                       stockout_days, stockout_date_calc, risk_level, reason
                FROM replenishment_decisions
                WHERE snapshot_id IN ({','.join(str(s) for s in existing_sids)})
            """)).fetchall()
            for row in existing_rows:
                existing_values[row.snapshot_id] = {
                    "summary_flag": str(row.summary_flag or ""),
                    "asin": str(row.asin or ""),
                    "sku": str(row.sku or ""),
                    "account": str(row.account or ""),
                    "country": str(row.country or ""),
                    "snapshot_date": row.snapshot_date,
                    "future_stock": row.future_stock,
                    "demand": row.demand,
                    "suggest_qty": row.suggest_qty,
                    "days_of_supply": row.days_of_supply,
                    "stockout_days": row.stockout_days,
                    "stockout_date_calc": str(row.stockout_date_calc or ""),
                    "risk_level": str(row.risk_level or ""),
                    "reason": str(row.reason or ""),
                }
        except Exception:
            existing_values = {}

    # 2. 收集需要写入的行（新记录 + 有变更的记录）
    write_batches = []
    for i in range(n):
        if i >= len(snapshot_ids):
            break
        sid = snapshot_ids[i]
        sf = str(summary_flags[i]) if summary_flags[i] else "0"
        stockout_date_str = "-" if stockout_dates[i] == "-" else stockout_dates[i].strftime('%Y-%m-%d')
        esc_asin = str(asins[i]).replace("'", "''")
        esc_sku = str(skus[i]).replace("'", "''")
        esc_acct = str(accounts[i]).replace("'", "''")
        esc_ctry = str(countries[i]).replace("'", "''")
        esc_reason = str(reasons[i]).replace("'", "''")

        new_fs = int(future_stock[i])
        new_dm = int(demand[i])
        new_sq = int(suggest_qty[i])
        new_dos = round(days_of_supply[i], 1)
        new_sod = int(stockout_days[i])

        # Diff 比较：已有记录且所有字段一致（含 snapshot_date）则跳过
        # 注意：snapshot_date 必须参与比较，否则旧 dec 记录会停留在上次计算的日期，
        # 导致 search_inventory 的 JOIN 条件 dec.snapshot_date == latest 匹配不上
        old = existing_values.get(sid)
        if old is not None:
            if (old["summary_flag"] == sf
                    and old["asin"] == esc_asin
                    and old["sku"] == esc_sku
                    and old["account"] == esc_acct
                    and old["country"] == esc_ctry
                    and old.get("snapshot_date") == target_date
                    and old["future_stock"] == new_fs
                    and old["demand"] == new_dm
                    and old["suggest_qty"] == new_sq
                    and old["days_of_supply"] == new_dos
                    and old["stockout_days"] == new_sod
                    and old["stockout_date_calc"] == stockout_date_str
                    and old["risk_level"] == risk_levels[i]
                    and old["reason"] == esc_reason):
                continue

        write_batches.append(
            f"(1,{sid},'{sf}','{esc_asin}','{esc_sku}','{esc_acct}','{esc_ctry}',"
            f"'{target_date}',{new_fs},{new_dm},0,{new_sq},"
            f"{new_dos},{new_sod},"
            f"'{stockout_date_str}','{risk_levels[i]}','{esc_reason}')"
        )

    # 3. 批量写入（INSERT ... ON DUPLICATE KEY UPDATE）
    if write_batches:
        BATCH_SIZE = 1000
        total_write = len(write_batches)
        for batch_start in range(0, total_write, BATCH_SIZE):
            batch = write_batches[batch_start:batch_start + BATCH_SIZE]
            sql = f"""
                INSERT INTO replenishment_decisions 
                (tenant_id, snapshot_id, summary_flag, asin, sku, account, country, snapshot_date, 
                 future_stock, demand, safety_stock, suggest_qty, days_of_supply, stockout_days, 
                 stockout_date_calc, risk_level, reason)
                VALUES {','.join(batch)}
                ON DUPLICATE KEY UPDATE
                    summary_flag=VALUES(summary_flag),
                    asin=VALUES(asin),
                    sku=VALUES(sku),
                    account=VALUES(account),
                    country=VALUES(country),
                    snapshot_date=VALUES(snapshot_date),
                    future_stock=VALUES(future_stock),
                    demand=VALUES(demand),
                    safety_stock=VALUES(safety_stock),
                    suggest_qty=VALUES(suggest_qty),
                    days_of_supply=VALUES(days_of_supply),
                    stockout_days=VALUES(stockout_days),
                    stockout_date_calc=VALUES(stockout_date_calc),
                    risk_level=VALUES(risk_level),
                    reason=VALUES(reason),
                    deleted_at=NULL
            """
            db.execute(text(sql))
            db.commit()

            if progress_callback:
                processed = min(batch_start + BATCH_SIZE, total_write)
                pct = 40 + int(processed / total_write * 30)
                progress_callback("写入数据库", pct, 100)

    if progress_callback:
        progress_callback("写入数据库", 100, 100)

    # 统计
    red = int(np.sum(risk_levels == "红"))
    yellow = int(np.sum(risk_levels == "黄"))
    green = int(np.sum(risk_levels == "绿"))

    return {
        "date": target_date.isoformat(),
        "total": n,
        "red": red,
        "yellow": yellow,
        "green": green,
    }


def calculate_replenishment(db: Session, snapshot_date: str = None, snapshot_ids: list = None, progress_callback=None) -> dict:
    """公开API：计算补货决策"""
    # 查找最新快照日期
    latest = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()

    if snapshot_date:
        target_date = datetime.strptime(snapshot_date, "%Y-%m-%d").date()
    elif latest:
        target_date = latest
    else:
        return {"message": "无快照数据", "date": None, "total": 0, "red": 0, "yellow": 0, "green": 0}

    # 获取快照
    base_filter = [
        InventorySnapshot.snapshot_date == target_date,
        InventorySnapshot.deleted_at.is_(None)
    ]
    if snapshot_ids:
        base_filter.append(InventorySnapshot.id.in_(snapshot_ids))
    snapshots = db.query(InventorySnapshot).filter(*base_filter).all()

    if not snapshots:
        return {"message": f"日期 {target_date} 无快照数据", "date": target_date.isoformat(), "total": 0, "red": 0, "yellow": 0, "green": 0}

    # 汇总行的本地仓库存需要累加子行（summary_flag="共享库存"）
    summary_asins = set(s.asin for s in snapshots if s.asin and s.summary_flag == "是")
    child_local_map = {}
    if summary_asins:
        child_rows = db.query(
            InventorySnapshot.asin,
            func.coalesce(func.sum(InventorySnapshot.local_inventory), 0).label("total_local")
        ).filter(
            InventorySnapshot.snapshot_date == target_date,
            InventorySnapshot.deleted_at.is_(None),
            InventorySnapshot.asin.in_(summary_asins),
            InventorySnapshot.summary_flag == "共享库存"
        ).group_by(InventorySnapshot.asin).all()
        child_local_map = {asin: total for asin, total in child_rows}

    # 转为DataFrame计算
    data = [{
        "id": s.id,
        "summary_flag": s.summary_flag or "0",
        "daily_sales": s.daily_sales or 0,
        "fba_stock": s.fba_stock or 0,
        "fba_inbound": s.fba_inbound or 0,
        "local_inventory": (s.local_inventory or 0) + float(child_local_map.get(s.asin, 0)) if s.summary_flag == "是" else (s.local_inventory or 0),
        "inspection_quantity": s.inspection_quantity or 0,
        "total_stock": s.total_stock or 0,
        "asin": s.asin or "",
        "sku": s.sku or "",
        "account": s.account or "",
        "country": s.country or "",
        # 添加 daily_avg 字段用于重新计算
        "daily_avg_3d": s.daily_avg_3d or 0,
        "daily_avg_7d": s.daily_avg_7d or 0,
        "daily_avg_14d": s.daily_avg_14d or 0,
        "daily_avg_30d": s.daily_avg_30d or 0,
        "daily_avg_60d": s.daily_avg_60d or 0,
        "daily_avg_90d": s.daily_avg_90d or 0,
    } for s in snapshots]

    df = pd.DataFrame(data)
    # 动态计算总库存：fba_stock + fba_inbound + local_inventory - inspection_quantity
    df["total_stock"] = (
        df["fba_stock"].fillna(0) +
        df["fba_inbound"].fillna(0) +
        df["local_inventory"].fillna(0) -
        df["inspection_quantity"].fillna(0)
    )
    snapshot_ids = df["id"].tolist()

    # 使用当前公式重新计算日均销量
    print("[DEBUG] 重新计算日均销量...")
    new_daily_sales = _calculate_daily_sales(df, db)
    df["daily_sales"] = new_daily_sales

    # 批量更新数据库中的日均销量
    print(f"[DEBUG] 批量更新 {len(snapshot_ids)} 条记录的日均销量...")
    try:
        # 使用 CASE WHEN 批量更新
        case_statements = []
        for i, sid in enumerate(snapshot_ids):
            case_statements.append(f"WHEN {sid} THEN {new_daily_sales[i]}")

        if case_statements:
            case_sql = " ".join(case_statements)
            ids_sql = ",".join([str(s) for s in snapshot_ids])
            update_sql = f"""
                UPDATE inventory_snapshots
                SET daily_sales = CASE id
                    {case_sql}
                END
                WHERE id IN ({ids_sql})
            """
            db.execute(text(update_sql))
            db.commit()
            print("[DEBUG] 日均销量更新完成")
    except Exception as e:
        print(f"[WARN] 批量更新日均销量失败: {e}")
        db.rollback()

    # 软删除由 UPSERT 自动处理，无需手动 DELETE

    return _calculate_replenishment_fast(db, df, snapshot_ids, target_date, progress_callback=progress_callback)


def get_inventory_overview(db: Session, user_id: int = None, user_role: str = None) -> dict:
    """获取库存概览"""
    latest = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()
    if not latest:
        return {"total_sku": 0, "red_count": 0, "yellow_count": 0, "green_count": 0, 
                "snapshot_date": None, "stockout_top10": [], "overstock_top10": []}

    # 构建基础查询（排除共享库存、节日产品和停售产品）
    base_query = db.query(InventorySnapshot).filter(
        InventorySnapshot.snapshot_date == latest,
        InventorySnapshot.deleted_at.is_(None),
        (InventorySnapshot.summary_flag != "共享库存") | (InventorySnapshot.summary_flag.is_(None)),
        # 排除节日产品（is_holiday != ""）
        (InventorySnapshot.is_holiday == "") | (InventorySnapshot.is_holiday.is_(None)),
        # 排除停售产品
        (InventorySnapshot.is_discontinued == False) | (InventorySnapshot.is_discontinued.is_(None)),
    )

    # 用户数据隔离
    if user_id and user_role and user_role != "admin":
        from sqlalchemy import text as sql_text
        dept_rows = db.execute(
            sql_text("SELECT department_id FROM user_departments WHERE user_id = :uid"),
            {"uid": user_id}
        ).fetchall()
        dept_ids = [d[0] for d in dept_rows if d[0]]
        if dept_ids:
            store_rows = db.execute(
                sql_text("""
                    SELECT inventory_name FROM stores
                    WHERE department_id IN :dept_ids
                    AND status = 'active'
                    AND inventory_name IS NOT NULL
                    AND inventory_name != ''
                """),
                {"dept_ids": tuple(dept_ids)}
            ).fetchall()
            user_stores = [s[0] for s in store_rows]
            if user_stores:
                store_conditions = _build_account_filter(user_stores)
                base_query = base_query.filter(or_(*store_conditions)) if store_conditions else base_query
            else:
                return {"total_sku": 0, "red_count": 0, "yellow_count": 0, "green_count": 0, 
                        "snapshot_date": latest.isoformat(), "stockout_top10": [], "overstock_top10": []}
        else:
            return {"total_sku": 0, "red_count": 0, "yellow_count": 0, "green_count": 0, 
                    "snapshot_date": latest.isoformat(), "stockout_top10": [], "overstock_top10": []}

    # 统计
    total = base_query.count()

    # 获取有效的 snapshot_ids（使用子查询，避免将 4 万条 ID 取回应用层）
    valid_snap_ids_subquery = base_query.with_entities(InventorySnapshot.id)

    # 用 LEFT JOIN 统计风险等级：无决策的快照计入"绿"，确保 red+yellow+green=total
    risk_counts_rows = db.query(
        func.coalesce(ReplenishmentDecision.risk_level, '绿'),
        func.count(InventorySnapshot.id)
    ).outerjoin(
        ReplenishmentDecision,
        and_(
            ReplenishmentDecision.snapshot_id == InventorySnapshot.id,
            ReplenishmentDecision.deleted_at.is_(None),
        )
    ).filter(
        InventorySnapshot.id.in_(valid_snap_ids_subquery)
    ).group_by(
        func.coalesce(ReplenishmentDecision.risk_level, '绿')
    ).all()
    risk_counts = {row[0]: row[1] for row in risk_counts_rows}
    red = risk_counts.get("红", 0)
    yellow = risk_counts.get("黄", 0)
    green = risk_counts.get("绿", 0)

    # TOP10
    stockout_ids = db.query(ReplenishmentDecision.snapshot_id).filter(
        ReplenishmentDecision.snapshot_date == latest,
        ReplenishmentDecision.deleted_at.is_(None),
        ReplenishmentDecision.snapshot_id.in_(valid_snap_ids_subquery),
    ).order_by(ReplenishmentDecision.days_of_supply.asc()).limit(10).all()
    snap_ids = [r[0] for r in stockout_ids]

    stockout_items = []
    if snap_ids:
        snaps = {s.id: s for s in db.query(InventorySnapshot).filter(InventorySnapshot.id.in_(snap_ids), InventorySnapshot.deleted_at.is_(None)).all()}
        decs = {d.snapshot_id: d for d in db.query(ReplenishmentDecision).filter(ReplenishmentDecision.snapshot_id.in_(snap_ids), ReplenishmentDecision.deleted_at.is_(None)).all()}
        for sid in snap_ids:
            if sid in snaps and sid in decs:
                s, d = snaps[sid], decs[sid]
                stockout_items.append({
                    "asin": s.asin, "product_name": s.product_name, "account": s.account,
                    "country": s.country, "days_of_supply": d.days_of_supply,
                    "fba_stock": s.fba_stock, "daily_sales": s.daily_sales,
                    "stockout_date": d.stockout_date_calc or "-",
                    "total_stock": (s.fba_stock or 0) + (s.fba_inbound or 0) + (s.local_inventory or 0) - (s.inspection_quantity or 0), "age_0_3": s.age_0_3, "age_3_6": s.age_3_6,
                    "age_6_9": s.age_6_9, "age_9_12": s.age_9_12, "age_12_plus": s.age_12_plus,
                    "suggest_qty": int(d.suggest_qty) if d.suggest_qty else 0,
                    "reason": d.reason or "",
                })

    overstock_top10 = []
    overstock_snaps = db.query(InventorySnapshot).filter(
        InventorySnapshot.snapshot_date == latest,
        InventorySnapshot.deleted_at.is_(None),
        InventorySnapshot.id.in_(valid_snap_ids_subquery),
    ).order_by(InventorySnapshot.age_12_plus.desc()).limit(10).all()
    if overstock_snaps:
        overstock_top10 = [{
            "asin": s.asin, "product_name": s.product_name, "account": s.account,
            "country": s.country, "total_stock": (s.fba_stock or 0) + (s.fba_inbound or 0) + (s.local_inventory or 0) - (s.inspection_quantity or 0),
            "age_0_3": s.age_0_3, "age_3_6": s.age_3_6, "age_6_9": s.age_6_9,
            "age_9_12": s.age_9_12, "age_12_plus": s.age_12_plus,
        } for s in overstock_snaps]

    return {
        "snapshot_date": latest.isoformat(),
        "total_sku": total,
        "red_count": red,
        "yellow_count": yellow,
        "green_count": green,
        "stockout_top10": stockout_items,
        "overstock_top10": overstock_top10,
    }


def search_inventory(db: Session, keyword: str = None, risk_level=None,
                     replenishment_status: str = None, account: str = None,
                     country: str = None, holiday_filter: str = None,
                     sort_field: str = None, sort_order: str = None,
                     page: int = 1, page_size: int = 20,
                     user_id: int = None, user_role: str = None) -> dict:
    """搜索库存"""
    # 用户数据隔离：获取用户可见的店铺列表
    user_stores = None
    if user_id and user_role and user_role != "admin":
        from sqlalchemy import text as sql_text
        dept_rows = db.execute(
            sql_text("SELECT department_id FROM user_departments WHERE user_id = :uid"),
            {"uid": user_id}
        ).fetchall()
        dept_ids = [d[0] for d in dept_rows if d[0]]
        if dept_ids:
            # 使用原生 SQL 查询避免模型导入问题
            store_rows = db.execute(
                sql_text("""
                    SELECT inventory_name FROM stores
                    WHERE department_id IN :dept_ids
                    AND status = 'active'
                    AND inventory_name IS NOT NULL
                    AND inventory_name != ''
                """),
                {"dept_ids": tuple(dept_ids)}
            ).fetchall()
            user_stores = [s[0] for s in store_rows]
        else:
            # 用户不属于任何部门，返回空
            return {"items": [], "total": 0, "page": page, "page_size": page_size}

    latest = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()
    if not latest:
        return {"items": [], "total": 0, "page": page, "page_size": page_size}

    # 1. 筛选要展示的 snapshot_ids
    #    - 普通行（summary_flag != '共享库存'）直接显示
    #    - 子行（summary_flag='共享库存'）只显示"孤儿子行"——即 ASIN 没有对应父汇总行（summary_flag='是'）的子行
    #      有父汇总行的子行通过父行展开按钮加载（get_summary_children API），主列表不重复显示
    valid_snapshot_query = db.query(InventorySnapshot.id).filter(
        InventorySnapshot.snapshot_date == latest,
        InventorySnapshot.deleted_at.is_(None),
        or_(
            InventorySnapshot.summary_flag != "共享库存",
            InventorySnapshot.summary_flag.is_(None),
            # 孤儿子行：summary_flag='共享库存' 但该 ASIN 在最新日期没有 summary_flag='是' 的父汇总行
            and_(
                InventorySnapshot.summary_flag == "共享库存",
                ~InventorySnapshot.asin.in_(
                    db.query(InventorySnapshot.asin).filter(
                        InventorySnapshot.snapshot_date == latest,
                        InventorySnapshot.deleted_at.is_(None),
                        InventorySnapshot.summary_flag == "是"
                    )
                )
            )
        )
    )

    # 应用用户数据隔离
    if user_stores is not None:
        if len(user_stores) == 0:
            # 用户有部门但没有关联店铺，返回空数据
            return {"items": [], "total": 0, "page": page, "page_size": page_size}
        # 使用精确匹配 + 汇总行边界匹配，支持多店铺集合数据（如 "D-EU-DE、D-EU-FR"）
        store_conditions = _build_account_filter(user_stores)
        if store_conditions:
            valid_snapshot_query = valid_snapshot_query.filter(or_(*store_conditions))

    if keyword:
        kw = f"%{keyword}%"
        valid_snapshot_query = valid_snapshot_query.filter(
            (InventorySnapshot.asin.like(kw)) |
            (InventorySnapshot.sku.like(kw)) |
            (InventorySnapshot.product_name.like(kw)) |
            (InventorySnapshot.account.like(kw))
        )

    # 节日产品筛选
    if holiday_filter and holiday_filter != "all":
        if holiday_filter == "only_non_holiday":
            # 正常产品：非节日且未停售
            valid_snapshot_query = valid_snapshot_query.filter(
                (InventorySnapshot.is_holiday == "") | (InventorySnapshot.is_holiday.is_(None))
            ).filter(
                (InventorySnapshot.is_discontinued == False) | (InventorySnapshot.is_discontinued.is_(None))
            )
        elif holiday_filter == "only_holiday":
            # 所有节日产品（未停售）
            valid_snapshot_query = valid_snapshot_query.filter(
                (InventorySnapshot.is_holiday != "") & (InventorySnapshot.is_holiday.isnot(None))
            ).filter(
                (InventorySnapshot.is_discontinued == False) | (InventorySnapshot.is_discontinued.is_(None))
            )
        elif holiday_filter == "only_discontinued":
            # 停售产品
            valid_snapshot_query = valid_snapshot_query.filter(
                InventorySnapshot.is_discontinued == True
            )
        elif holiday_filter.startswith("holiday_"):
            # 按具体节日类型筛选（如 holiday_圣诞 / holiday_万圣 / holiday_跨年 / holiday_其他）
            holiday_type = holiday_filter[len("holiday_"):]
            valid_snapshot_query = valid_snapshot_query.filter(
                InventorySnapshot.is_holiday == holiday_type
            )

    # 国家和店铺筛选逻辑：
    # 1. 如果只选国家：筛选该国家的所有库存（基于店铺字段匹配）
    # 2. 如果只选店铺：筛选包含该店铺的所有库存
    # 3. 如果同时选国家和店铺：以店铺筛选为准（店铺已包含国家信息）

    if account:
        # account 参数是 "店铺-国家" 格式（如 "JeVenis-US"）
        # 使用精确匹配 + 汇总行边界匹配，避免 'G-UK' 误匹配 'Roaring-UK-UK'
        account_conditions = _build_account_filter(account)
        if account_conditions:
            valid_snapshot_query = valid_snapshot_query.filter(or_(*account_conditions))
    elif country:
        # 只选了国家，没选店铺：基于 country 字段筛选
        # 将国家名称转换为站点代码进行匹配
        country_to_site_map = {
            '美国': ['美国', '美国站', 'US', 'USA'],
            '英国': ['英国', 'UK'],
            '德国': ['德国', 'DE'],
            '法国': ['法国', 'FR'],
            '意大利': ['意大利', 'IT'],
            '西班牙': ['西班牙', 'ES'],
            '日本': ['日本', 'JP'],
            '加拿大': ['加拿大', 'CA'],
            '墨西哥': ['墨西哥', 'MX'],
            '澳大利亚': ['澳大利亚', 'AU'],
            '荷兰': ['荷兰', 'NL'],
            '瑞典': ['瑞典', 'SE'],
            '波兰': ['波兰', 'PL'],
            '比利时': ['比利时', 'BE'],
            '爱尔兰': ['爱尔兰', 'IE'],
            '新加坡': ['新加坡', 'SG'],
            '阿联酋': ['阿联酋', 'AE'],
            '印度': ['印度', 'IN'],
            '巴西': ['巴西', 'BR'],
            '土耳其': ['土耳其', 'TR'],
        }
        
        # 构建国家筛选条件：匹配 country 字段
        country_conditions = []
        country_list = [country] if isinstance(country, str) else country
        for c in country_list:
            site_codes = country_to_site_map.get(c, [c])
            for code in site_codes:
                country_conditions.append(InventorySnapshot.country.like(f"%{code}%"))
        
        if country_conditions:
            valid_snapshot_query = valid_snapshot_query.filter(or_(*country_conditions))

    # 改用子查询，避免把 4 万条 ID 取回应用层
    valid_snapshot_subquery = valid_snapshot_query

    # 3. 构建主查询，使用子查询筛选
    query = db.query(InventorySnapshot, ReplenishmentDecision).outerjoin(
        ReplenishmentDecision,
        (ReplenishmentDecision.snapshot_id == InventorySnapshot.id) &
        (ReplenishmentDecision.snapshot_date == latest)
    ).filter(
        InventorySnapshot.id.in_(valid_snapshot_subquery),
        InventorySnapshot.deleted_at.is_(None)
    )

    # 如果有风险等级筛选，添加 WHERE 条件（outerjoin + WHERE 会自然过滤掉无匹配的行）
    if risk_level:
        risk_map = {"red": "红", "yellow": "黄", "green": "绿"}
        rls = [risk_map.get(rl, rl) for rl in ([risk_level] if isinstance(risk_level, str) else risk_level)]
        query = query.filter(ReplenishmentDecision.risk_level.in_(rls))

    if replenishment_status:
        query = query.filter(InventorySnapshot.replenishment_status == replenishment_status)

    if sort_field == 'days_of_supply':
        order_col = ReplenishmentDecision.days_of_supply.desc() if sort_order == 'desc' else ReplenishmentDecision.days_of_supply.asc()
        query = query.order_by(order_col, InventorySnapshot.asin)
    elif sort_field == 'suggest_qty':
        order_col = ReplenishmentDecision.suggest_qty.desc() if sort_order == 'desc' else ReplenishmentDecision.suggest_qty.asc()
        query = query.order_by(order_col, InventorySnapshot.asin)
    elif sort_field == 'stockout_date':
        # 断货时间前端显示取自 ReplenishmentDecision.stockout_date_calc（字符串 YYYY-MM-DD）
        # 需用相同字段排序，并用 CASE 将无效值（NULL/'-'/'') 统一处理，避免与显示不一致
        # 升序时无效值替换为最大值 '9999-12-31'（排到最后）
        # 降序时无效值替换为最小值 '0001-01-01'（排到最后）
        from sqlalchemy import case, literal_column
        invalid = or_(
            ReplenishmentDecision.stockout_date_calc.is_(None),
            ReplenishmentDecision.stockout_date_calc == "",
            ReplenishmentDecision.stockout_date_calc == "-",
        )
        placeholder = "'0001-01-01'" if sort_order == 'desc' else "'9999-12-31'"
        sort_expr = case(
            (invalid, literal_column(placeholder)),
            else_=ReplenishmentDecision.stockout_date_calc
        )
        query = query.order_by(
            sort_expr.desc() if sort_order == 'desc' else sort_expr.asc(),
            InventorySnapshot.asin
        )
    elif sort_field:
        col = getattr(InventorySnapshot, sort_field, None)
        if col is not None:
            query = query.order_by(col.desc() if sort_order == 'desc' else col.asc())
    else:
        query = query.order_by(InventorySnapshot.asin)

    total = query.count()
    results = query.offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for snap, dec in results:
        risk_map_r = {"红": "red", "黄": "yellow", "绿": "green"}
        local_qty = snap.local_inventory or 0
        items.append({
            "id": snap.id, "asin": snap.asin, "sku": snap.sku, "fnsku": snap.fnsku,
            "msku": snap.msku, "product_name": snap.product_name, "account": snap.account,
            "country": snap.country, "category": snap.category, "brand": snap.brand,
            "replenishment_status": snap.replenishment_status,
            "days_of_supply": dec.days_of_supply if dec else snap.days_supply_total,
            "fba_stock": snap.fba_stock, "fba_available": snap.fba_available,
            "fba_pending_transfer": snap.fba_pending_transfer, "fba_in_transfer": snap.fba_in_transfer,
            "fba_inbound_processing": snap.fba_inbound_processing, "fba_inbound": snap.fba_inbound,
            "daily_sales": round(snap.daily_sales, 2) if snap.daily_sales is not None else None, "total_stock": (snap.fba_stock or 0) + (snap.fba_inbound or 0) + (snap.local_inventory or 0) - (snap.inspection_quantity or 0),
            "sales_7d": snap.sales_7d,
            "sales_14d": snap.sales_14d,
            "daily_avg_7d": round(snap.daily_avg_7d, 2) if snap.daily_avg_7d else None,
            "daily_avg_14d": round(snap.daily_avg_14d, 2) if snap.daily_avg_14d else None,
            "stockout_date": dec.stockout_date_calc if dec else (snap.stockout_date.isoformat() if snap.stockout_date else "-"),
            "age_12_plus": snap.age_12_plus,
            "risk_level": risk_map_r.get(dec.risk_level, "green") if dec else "green",
            "summary_flag": snap.summary_flag,
            "local_inventory": local_qty,
            "suggest_qty": int(dec.suggest_qty) if dec else 0,
            "replenishment_reason": dec.reason if dec else "",
            "inspection_quantity": snap.inspection_quantity or 0,
            "gross_margin": snap.gross_margin,
            "is_holiday": snap.is_holiday or "",
            "is_discontinued": snap.is_discontinued if snap.is_discontinued is not None else False,
        })

    # 汇总行的 local_inventory = 子行 local_inventory 之和
    summary_asins = [item["asin"] for item in items if item["summary_flag"] == "是"]
    if summary_asins:
        child_sums = db.query(
            InventorySnapshot.asin,
            func.coalesce(func.sum(InventorySnapshot.local_inventory), 0)
        ).filter(
            InventorySnapshot.snapshot_date == latest,
            InventorySnapshot.deleted_at.is_(None),
            InventorySnapshot.asin.in_(summary_asins),
            InventorySnapshot.summary_flag == "共享库存"
        ).group_by(InventorySnapshot.asin).all()
        sum_map = {asin: total for asin, total in child_sums}
        for item in items:
            if item["summary_flag"] == "是":
                child_local = sum_map.get(item["asin"], 0)
                item["local_inventory"] = child_local
                item["total_stock"] = float(item.get("fba_stock") or 0) + float(item.get("fba_inbound") or 0) + float(child_local) - float(item.get("inspection_quantity") or 0)

    return {"items": items, "total": total, "page": page, "page_size": page_size}


def get_stockout_top10(db: Session, user_id: int = None, user_role: str = None) -> list:
    """断货风险TOP10"""
    latest = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()
    if not latest:
        return []

    # 用户数据隔离：先计算可见的 snapshot_ids
    visible_snap_ids = None
    if user_id and user_role and user_role != "admin":
        from sqlalchemy import text as sql_text
        dept_rows = db.execute(
            sql_text("SELECT department_id FROM user_departments WHERE user_id = :uid"),
            {"uid": user_id}
        ).fetchall()
        dept_ids = [d[0] for d in dept_rows if d[0]]
        if not dept_ids:
            return []
        store_rows = db.execute(
            sql_text("""
                SELECT inventory_name FROM stores
                WHERE department_id IN :dept_ids
                AND status = 'active'
                AND inventory_name IS NOT NULL
                AND inventory_name != ''
            """),
            {"dept_ids": tuple(dept_ids)}
        ).fetchall()
        user_stores = [s[0] for s in store_rows]
        if not user_stores:
            return []
        # 获取可见的 snapshot_ids
        snap_store_conditions = _build_account_filter(user_stores)
        visible_snap_ids = [s.id for s in db.query(InventorySnapshot.id).filter(
            InventorySnapshot.snapshot_date == latest,
            InventorySnapshot.deleted_at.is_(None),
            or_(*snap_store_conditions)
        ).all()] if snap_store_conditions else []
        if not visible_snap_ids:
            return []

    # 单条 JOIN 查询合并原 3 次查询（snapshot_id 查询 + snaps 查询 + decs 查询）
    join_query = db.query(InventorySnapshot, ReplenishmentDecision).join(
        ReplenishmentDecision,
        and_(
            ReplenishmentDecision.snapshot_id == InventorySnapshot.id,
            ReplenishmentDecision.snapshot_date == latest,
            ReplenishmentDecision.deleted_at.is_(None),
            (ReplenishmentDecision.summary_flag != "共享库存") | (ReplenishmentDecision.summary_flag.is_(None))
        )
    ).filter(
        InventorySnapshot.deleted_at.is_(None),
        # 排除节日产品和停售产品
        (InventorySnapshot.is_holiday == "") | (InventorySnapshot.is_holiday.is_(None)),
        (InventorySnapshot.is_discontinued == False) | (InventorySnapshot.is_discontinued.is_(None)),
    )

    if visible_snap_ids is not None:
        join_query = join_query.filter(InventorySnapshot.id.in_(visible_snap_ids))

    results = join_query.order_by(ReplenishmentDecision.days_of_supply.asc()).limit(10).all()

    if not results:
        return []

    # 直接遍历 JOIN 结果组装数据
    result = []
    for snap, dec in results:
        result.append({
            "asin": snap.asin, "product_name": snap.product_name, "account": snap.account,
            "country": snap.country, "days_of_supply": dec.days_of_supply,
            "fba_stock": snap.fba_stock, "daily_sales": snap.daily_sales,
            "stockout_date": dec.stockout_date_calc or "-",
            "total_stock": (snap.fba_stock or 0) + (snap.fba_inbound or 0) + (snap.local_inventory or 0) - (snap.inspection_quantity or 0),
            "age_0_3": snap.age_0_3, "age_3_6": snap.age_3_6,
            "age_6_9": snap.age_6_9, "age_9_12": snap.age_9_12, "age_12_plus": snap.age_12_plus,
            "suggest_qty": int(dec.suggest_qty) if dec.suggest_qty else 0,
            "reason": dec.reason or "",
        })
    return result


def get_overstock_top10(db: Session, user_id: int = None, user_role: str = None) -> list:
    """冗余库存TOP10"""
    latest = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()
    if not latest:
        return []

    base_query = db.query(InventorySnapshot).filter(
        InventorySnapshot.snapshot_date == latest,
        InventorySnapshot.deleted_at.is_(None),
        (InventorySnapshot.summary_flag != "共享库存") | (InventorySnapshot.summary_flag.is_(None)),
        # 排除节日产品和停售产品
        (InventorySnapshot.is_holiday == "") | (InventorySnapshot.is_holiday.is_(None)),
        (InventorySnapshot.is_discontinued == False) | (InventorySnapshot.is_discontinued.is_(None)),
    )

    # 用户数据隔离
    if user_id and user_role and user_role != "admin":
        from sqlalchemy import text as sql_text
        dept_rows = db.execute(
            sql_text("SELECT department_id FROM user_departments WHERE user_id = :uid"),
            {"uid": user_id}
        ).fetchall()
        dept_ids = [d[0] for d in dept_rows if d[0]]
        if dept_ids:
            store_rows = db.execute(
                sql_text("""
                    SELECT inventory_name FROM stores
                    WHERE department_id IN :dept_ids
                    AND status = 'active'
                    AND inventory_name IS NOT NULL
                    AND inventory_name != ''
                """),
                {"dept_ids": tuple(dept_ids)}
            ).fetchall()
            user_stores = [s[0] for s in store_rows]
            if not user_stores:
                return []
            store_conditions = _build_account_filter(user_stores)
            if store_conditions:
                base_query = base_query.filter(or_(*store_conditions))
        else:
            return []

    items = base_query.order_by(InventorySnapshot.age_12_plus.desc()).limit(10).all()

    return [{
        "asin": s.asin, "product_name": s.product_name, "account": s.account,
        "country": s.country, "total_stock": (s.fba_stock or 0) + (s.fba_inbound or 0) + (s.local_inventory or 0) - (s.inspection_quantity or 0),
        "age_0_3": s.age_0_3, "age_3_6": s.age_3_6, "age_6_9": s.age_6_9,
        "age_9_12": s.age_9_12, "age_12_plus": s.age_12_plus,
    } for s in items]


def get_inbound_details(db: Session, asin: str, account: str = None) -> list:
    """查询在途详情"""
    latest = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()
    if not latest:
        return []

    query = db.query(InventorySnapshot.id).filter(
        InventorySnapshot.snapshot_date == latest, InventorySnapshot.asin == asin,
        InventorySnapshot.deleted_at.is_(None)
    )
    if account:
        query = query.filter(InventorySnapshot.account == account)
    
    snapshot_ids = [s.id for s in query.all()]
    if not snapshot_ids:
        return []

    details = db.query(InboundShipmentDetail).filter(
        InboundShipmentDetail.snapshot_id.in_(snapshot_ids),
        InboundShipmentDetail.deleted_at.is_(None)
    ).all()

    return [{
        "shipment_id": d.shipment_id, "quantity": d.quantity,
        "transport_method": d.transport_method,
        "ship_date": d.ship_date.isoformat() if d.ship_date else None,
        "estimated_available_date": d.estimated_available_date.isoformat() if d.estimated_available_date else None,
        "estimated_arrival_date": d.estimated_arrival_date.isoformat() if d.estimated_arrival_date else None,
        "raw_text": d.raw_text,
    } for d in details]


def get_latest_snapshot_date(db: Session) -> str:
    """获取最新快照日期"""
    latest = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()
    return latest.isoformat() if latest else None


def get_summary_children(db: Session, asin: str, parent_account: str = None):
    """获取汇总行的子行数据

    Args:
        asin: 父汇总行的 ASIN
        parent_account: 父汇总行的 account 字段（如 "LaVenty--US、LaVenty--MX、LaVenty-US-BR"）。
            传入时会按 account 集合过滤子行，避免同一 ASIN 有多个汇总行时子行串扰
            （例：B07WHGXBNB 同时有北美3国汇总行 + 欧洲9国汇总行，展开北美父行只应看到北美3国子行）。
            不传则返回该 ASIN 所有 summary_flag='共享库存' 的子行（向后兼容）。
    """
    latest = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()
    if not latest:
        return []

    query = db.query(InventorySnapshot).filter(
        InventorySnapshot.snapshot_date == latest,
        InventorySnapshot.deleted_at.is_(None),
        InventorySnapshot.asin == asin,
        InventorySnapshot.summary_flag == "共享库存"
    )

    # 按父汇总行的 account 列表过滤子行
    if parent_account:
        # 父汇总行 account 形如 "A、B、C"，用顿号分割
        parent_accounts = [a.strip() for a in str(parent_account).split("、") if a.strip()]
        if parent_accounts:
            query = query.filter(InventorySnapshot.account.in_(parent_accounts))

    children = query.order_by(InventorySnapshot.account, InventorySnapshot.country).all()

    return [{
        "id": c.id,
        "asin": c.asin,
        "sku": c.sku,
        "fnsku": c.fnsku,
        "msku": c.msku,
        "product_name": c.product_name,
        "account": c.account,
        "country": c.country,
        "category": c.category,
        "brand": c.brand,
        "fba_stock": c.fba_stock,
        "fba_available": c.fba_available,
        "fba_pending_transfer": c.fba_pending_transfer,
        "fba_in_transfer": c.fba_in_transfer,
        "fba_inbound_processing": c.fba_inbound_processing,
        "fba_inbound": c.fba_inbound,
        "total_stock": (c.fba_stock or 0) + (c.fba_inbound or 0) + (c.local_inventory or 0) - (c.inspection_quantity or 0),
        "daily_sales": round(c.daily_sales, 2) if c.daily_sales is not None else None,
        "sales_7d": c.sales_7d,
        "sales_14d": c.sales_14d,
        "days_of_supply": c.days_supply_total,
        "stockout_date": c.stockout_date.isoformat() if c.stockout_date else None,
        "risk_level": None,
        "summary_flag": c.summary_flag,
        "suggest_qty": None,
        "replenishment_reason": None,
        "inspection_quantity": c.inspection_quantity or 0,
        "local_inventory": c.local_inventory or 0,
        "gross_margin": c.gross_margin,
    } for c in children]


def export_inventory_to_excel(items: list, fields: list = None):
    """将库存数据导出为Excel
    
    Args:
        items: 库存数据列表
        fields: 要导出的字段列表
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存明细"

    # 定义所有可选字段
    all_fields = {
        "asin": "ASIN",
        "sku": "SKU",
        "fnsku": "FNSKU",
        "msku": "MSKU",
        "product_name": "品名",
        "account": "店铺",
        "country": "国家",
        "category": "分类",
        "brand": "品牌",
        "fba_stock": "FBA库存",
        "fba_available": "可售",
        "fba_pending_transfer": "待调仓",
        "fba_in_transfer": "FBA预留",
        "fba_inbound_processing": "入库中",
        "fba_inbound": "在途",
        "local_inventory": "本地仓库存",
        "total_stock": "总库存",
        "gross_margin": "毛利率",
        "is_holiday": "节日类型",
        "is_discontinued": "停售",
        "daily_sales": "日均销量",
        "sales_7d": "7天销量",
        "sales_14d": "14天销量",
        "daily_avg_7d": "7日日均销量",
        "daily_avg_14d": "14日日均销量",
        "days_of_supply": "可售天数",
        "stockout_date": "断货时间",
        "risk_level": "风险等级",
        "suggest_qty": "建议补货数量",
        "replenishment_reason": "补货原因",
        "replenishment_status": "补货状态",
    }

    # 风险等级中文映射
    risk_map = {"red": "红", "yellow": "黄", "green": "绿"}

    # 确定导出字段
    if fields:
        export_fields = [(f, all_fields.get(f, f)) for f in fields if f in all_fields]
    else:
        export_fields = list(all_fields.items())

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 写表头
    for col, (key, label) in enumerate(export_fields, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 写数据
    for row_idx, item in enumerate(items, 2):
        for col, (key, _) in enumerate(export_fields, 1):
            val = item.get(key, "")
            if key == "risk_level":
                val = risk_map.get(val, val)
            elif key == "is_holiday":
                # 节日类型直接输出字符串值（已是中文）
                val = val if val else ""
            elif key == "is_discontinued":
                # 停售标记：True → "是"，False → "否"
                val = "是" if val else "否"
            elif key == "gross_margin":
                if val is not None and val != "":
                    try:
                        val = f"{float(val) * 100:.2f}%"
                    except (ValueError, TypeError):
                        pass
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # 添加自动筛选功能（在表头行）
    if export_fields:
        last_col_letter = openpyxl.utils.get_column_letter(len(export_fields))
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(items) + 1}"
        
        # 找到"建议补货数量"列并设置筛选条件（排除0值）
        suggest_qty_col_idx = None
        for idx, (key, _) in enumerate(export_fields, 1):
            if key == "suggest_qty":
                suggest_qty_col_idx = idx
                break
        
        if suggest_qty_col_idx:
            from openpyxl.worksheet.filters import FilterColumn, Filters
            
            # 收集所有非0的建议补货数量值
            non_zero_values = []
            for item in items:
                qty = item.get("suggest_qty") or 0
                if qty > 0:
                    val_str = str(int(qty))
                    if val_str not in non_zero_values:
                        non_zero_values.append(val_str)
            
            # 创建筛选：只显示非0值
            filters = Filters()
            filters.filter = non_zero_values  # 使用值列表而不是CustomFilter
            
            fc = FilterColumn(colId=suggest_qty_col_idx - 1, filters=filters)
            ws.auto_filter.filterColumn.append(fc)
            
            # 隐藏建议补货数量为0的行（Excel打开时直接显示筛选后的结果）
            for row_idx, item in enumerate(items, 2):  # 从第2行开始（第1行是表头）
                if (item.get("suggest_qty") or 0) == 0:
                    ws.row_dimensions[row_idx].hidden = True

    # 自动列宽
    for col in range(1, len(export_fields) + 1):
        max_len = 0
        for row in range(1, min(len(items) + 2, 100)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
